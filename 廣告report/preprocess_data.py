# coding: utf-8
""" 廣告成效周報 。整理各家代理商的成效報告。並整合後再按邏輯計算每日成效。

Version 1.0.0
Created on 2022/01/14 by Jerry.Ko
"""

import os
import datetime
import numpy as np
import pandas as pd
from dateutil.parser import parse
from openpyxl import load_workbook

def get_sheet(file_path:str):
    """處理各家代理商報表，只留下所需的欄位及資料。要求報表規則如下:
    A1為媒體名稱，A2為版位名稱。
    第3列(row)需為欄位名稱，其中A3必須為Date(日期)。
    資料表末端不能有加總列(row)，且位於結束時，需有一列空白欄做為提示。
    若讀取資料有問題時可以來這邊檢查。
    Args:
      file_path: str. 檔案路徑。
    Return:
      df: pandas dataframe. 帶有 media, banner, Date, Views, IMPs, Clicks 等欄位的表。
    """
    def parse_date(dt_date):
        """解析日期格式，處理str, int, datetime, date等格式轉換成datetime型別
        若代理商有特別的日期格式就來這裡改。
        Args:
          dt_date: str|int|float|datetime|date. 代理商傳遞進來的日期。
        """
        if isinstance(dt_date, str):
            return tmp_df.Date.apply(lambda d: parse(d))
        elif isinstance(dt_date, int) | isinstance(dt_date, float):
            return tmp_df.Date.apply(lambda d: parse(str(d)))
        elif isinstance(dt_date, datetime.datetime):
            return tmp_df.Date
        elif isinstance(dt_date, datetime.date):
            return pd.to_datetime(tmp_df.Date)

    workbook = load_workbook(file_path, data_only=True)
    sheetnames = workbook.sheetnames
    df = pd.DataFrame()
    for sheetname in sheetnames:
        ws = workbook[sheetname]

        media = ws['A1'].value
        banner = ws['A2'].value
        ws['A3'] = 'Date'
        # 紀錄完媒體、版位後，就不需要前兩行
        ws.delete_rows(1, 2)

        # 抓取資料範圍。idx就是開始沒有資料的row
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row[0]:
                date_data = row[0]
            else:
                idx -= 1
                break

        data = ws.values
        colnames = next(data)[0:]
        data = list(data)
        base_colname = ['Date', 'IMPs', 'Views', 'Clicks']
        # 有些檔案不會有 click 或 view欄位，因此需要用交集計算
        cols = set(base_colname).intersection(set(colnames))

        # 選取資料表的範圍以及所需存有欄位，並補齊所需欄位
        tmp_df = pd.DataFrame(data, columns=colnames).loc[0:idx-1, cols]
        diff_cols = set(base_colname).difference(set(tmp_df.columns))
        for i in diff_cols:
            tmp_df[i] = np.nan

        # 刪除有日期但無資料的筆數
        tmp_df = tmp_df[~(tmp_df[['IMPs', 'Clicks', 'Views']].sum(axis=1) == 0)]
        # 將 IMPs, clicks, views 改為 float，解析日期
        tmp_df[['IMPs', 'Clicks', 'Views']] = tmp_df[['IMPs', 'Clicks', 'Views']].astype('float')
        tmp_df['Date'] = parse_date(date_data)
        tmp_df['media'] = media
        tmp_df['banner'] = banner

        tmp_df = tmp_df.groupby(['media', 'banner', 'Date'], as_index=False).sum(min_count=1)

        df = df.append(tmp_df, ignore_index=True)
    return df

def preprocess_data(dir_path:str):
    """整合代理商的成效報告，並按邏輯計算每日成效。
    計算邏輯有誤時，可來這邊檢查。
    Args:
      dir_path: str. 資料夾路徑，內含所有代理商的檔案。
    """
    df_bk = pd.DataFrame()
    # 'D:/Project/ad_weekly_report/tmp'
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            file_path = os.path.join(root, file)
            print(file_path)
            tmp_df = get_sheet(file_path)
            df_bk = df_bk.append(tmp_df, ignore_index=True)
            
    df = df_bk.copy()
    
    #### 處理 AD-Daily總表 START ####
    
    # 計算媒體各版位每日加總，以及各媒體每日加總
    media_total = df.groupby(['media', 'Date'], as_index=False).sum(min_count=1)
    overall_total = df.groupby(['Date'], as_index=False).sum(min_count=1)

    df = df.append(media_total)
    df['banner'] = df.banner.fillna("加總")

    df = df.append(overall_total)
    df['media'] = df.media.fillna('總計Total')
    df['banner'] = df.banner.fillna('')

    # 計算各媒體版位不分日期加總
    date_total = df.groupby(['media', 'banner'], as_index=False, dropna=False).sum(min_count=1)
    df = df.append(date_total)
    
    # 要先算完加總後再算 duration date 才會有總計的走期
    duration_days = df.groupby(['media', 'banner'], dropna=False).agg(
        days = pd.NamedAgg(column="Date", aggfunc="nunique"),
        st_date = pd.NamedAgg(column="Date", aggfunc="min"),
        ed_date = pd.NamedAgg(column="Date", aggfunc="max"),
    )
    duration_days['走期'] = (  
           duration_days.st_date.dt.month.astype('str').str.zfill(2)
         + '/'
         + duration_days.st_date.dt.day.astype('str').str.zfill(2)
         + '-'
         + duration_days.ed_date.dt.month.astype('str').str.zfill(2)
         + '/'  
         + duration_days.ed_date.dt.day.astype('str').str.zfill(2)
    )
    duration_days = duration_days.set_index('走期', append=True)
    
    df['Date'] = df['Date'].dt.date
    df['Date'] = df.Date.fillna('Total')
    df['CTR'] = df.Clicks / df.IMPs
    df['VTR'] = df.Views / df.IMPs
    
    # 設定index 順序
    idx_lv1 = list(df.media.drop_duplicates().sort_values())
    idx_lv2 = list(df.banner.drop_duplicates().sort_values())
    # 將加總移至到list的最後
    idx_lv1.remove('總計Total')
    idx_lv1.append('總計Total')
    idx_lv2.remove('加總')
    idx_lv2.append('加總')
    
    final = df.melt(id_vars=['media', 'banner', 'Date'], value_vars=['IMPs', 'Clicks', 'CTR', 'Views', 'VTR'])\
        .pivot(['media', 'banner', 'variable'], 'Date', 'value')
    # mutliIndex 排序
    final = duration_days[['days']].join(final).drop('days', axis=1)\
        .reindex(idx_lv1, level='media')\
        .reindex(idx_lv2, level='banner')\
        .reindex(['IMPs', 'Clicks', 'CTR', 'Views', 'VTR'], level='variable')
    final.index.set_names('', level=3, inplace=True)
    #### 處理 AD-Daily總表 END ####
    
    #### 處理 AD事後評估總表 START ####
    date_total.index = pd.MultiIndex.from_frame(date_total[['media', 'banner']])
    performance_df = date_total.drop(['media', 'banner'], axis=1)

    performance_df['預估曝光數'] = np.nan
    performance_df['預估CPM'] = np.nan
    performance_df['實際CPM'] = ''
    performance_df['預估點閱次數'] = np.nan
    performance_df['預估CPC'] = np.nan
    performance_df['實際CPC'] = ''
    performance_df['預估點閱率'] = np.nan
    performance_df['實際總點閱率'] = performance_df.Clicks / performance_df.IMPs 
    performance_df['預估觀看次數'] = np.nan
    performance_df['預估CPV'] = np.nan
    performance_df['實際CPV'] = ''
    performance_df['預估觀看率'] = np.nan
    performance_df['實際總觀看率'] = performance_df.Views / performance_df.IMPs 
    performance_df['總預算$'] = np.nan
    performance_df['達標%'] = ''

    performance_df = duration_days[['days']].join(performance_df)\
        .reindex(idx_lv1, level='media')\
        .reindex(idx_lv2, level='banner')
    performance_df = performance_df[[
        'days', '預估曝光數', 'IMPs', '預估CPM', '實際CPM',
        '預估點閱次數', 'Clicks', '預估CPC', '實際CPC', '預估點閱率', '實際總點閱率',
        '預估觀看次數', 'Views', '預估CPV', '實際CPV', '預估觀看率', '實際總觀看率',
        '總預算$', '達標%'
    ]]
    #### 處理AD事後評估總表 END ####
    
    with pd.ExcelWriter(os.path.join(dir_path, 'output.xlsx')) as writer:
        final.to_excel(writer, sheet_name='AD-Daily總表')
        performance_df.to_excel(writer, sheet_name='AD事後評估總表')