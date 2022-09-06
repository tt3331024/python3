# coding: utf-8
""" 廣告成效周報 。
處理 excel 的 style。

Version 1.0.0
Created on 2022/01/14 by Jerry.Ko
"""
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def set_style_for_overview(file_path:str):
    """使用openpyxl設定style。最後不會回傳return，直接寫出檔案。
    總表的格式 style 都來這裡改。
    Args: 
      file_path: str. 檔案路徑。
    """
    workbook = load_workbook(file_path, data_only=True)
    ws = workbook[workbook.get_sheet_names()[0]]

    ws['A1'].value = '網站名稱\nWebsite'
    ws['B1'].value = '刊登版面\nPosition'
    ws['C1'].value = '走期\nPeriod'

    # 格式設定
    format_percentage = '0.00%;0.00%;"";@'
    format_0_decimal = '#,##0;#,##0;"";@'

    full_font = Font(name='微軟正黑體', size=12)
    title_font = Font(b=True)
    title_alignment = Alignment(horizontal='center', vertical='center')

    med_border = Border(bottom=Side(border_style='medium', color='000000'))
    title_med_border = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='medium', color='000000'))

    thin_border = Border(left=Side(border_style='dotted', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))
    title_thin_border = Border(left=Side(border_style='thin', color='000000'),
                               right=Side(border_style='thin', color='000000'),
                               top=Side(border_style='thin', color='000000'),
                               bottom=Side(border_style='thin', color='000000'))

    double_border = Border(left=Side(border_style='dotted', color='000000'),
                           bottom=Side(border_style='double', color='000000'))
    title_double_border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='double', color='000000'))

    date_fill = PatternFill(fill_type='solid', start_color='E0E0E0', end_color='E0E0E0')
    bg_fill = PatternFill(fill_type='solid', start_color='FFEEDD', end_color='FFEEDD')


    # 欄位設定: 寬與高
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 30.0

    for col in ['C', 'D']:
        ws.column_dimensions[col].width = 20.0

    for col in range(5, ws.max_column):
        ws.column_dimensions[get_column_letter(col)].width = 13.25

    ws.column_dimensions[get_column_letter(ws.max_column)].width = 17

    ws.row_dimensions[1].height = 40.0
    for row in range(2, ws.max_row+1):
        ws.row_dimensions[ws.max_row].height = 20.0

    # title格式設定: 儲存格置中/右 + 粗體
    for col in ['A', 'B', 'C']:
        for cell in ws[col]:
            cell.alignment = title_alignment
            cell.font = title_font

    for cell in ws['D']:
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = title_font

    for cell in ws[1]:
        cell.alignment = title_alignment
        cell.font = title_font

    # 全表設定: 字型、字體大小轉換
    for row in ws.iter_rows():
        for cell in row:
            cell.font = full_font

    # 資料格式設定: 數值格式轉換，小數點或加上千分位逗點
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            # 首欄為日期，不需format
            continue
        if (row[3].value == 'CTR') or (row[3].value == 'VTR'):
            for cell in row[4:]:
                cell.number_format = format_percentage
        else:
            for cell in row[4:]:
                cell.number_format = format_0_decimal

    # 設定框線及底色
    flag = False
    for i, row in enumerate(ws.iter_rows()):
        if row[1].value == '加總':
            flag = True
        if i == 0:
            for idx, cell in enumerate(row):
                if idx < 4:
                    cell.border = title_med_border
                else:
    #                 cell.fill = date_fill
                    cell.number_format = 'mm-dd'
                    cell.border = med_border
        elif i % 5 == 0:
            for idx, cell in enumerate(row):
                if idx < 4:
    #                 cell.border = title_thin_border
                    cell.border = title_double_border
                else:
    #                 cell.border = thin_border
                    cell.border = double_border
        else:
            for idx, cell in enumerate(row):
                if idx < 4:
                    pass
                else:
                    cell.border = thin_border

        # 當banner是'加總'時，會進到條件式中
        if flag:
            for idx, cell in enumerate(row):
                cell.fill = bg_fill
                if i % 5 == 0:
                    # 進到加總的最後一row後，關閉flag
                    flag = False
                    for idx, cell in enumerate(row):
                        if idx < 4:
                            cell.border = title_double_border
                        else:
                            cell.border = double_border
    # 加總Totla 處理                        
    for i, row in enumerate(ws[ws.max_row-4:ws.max_row]):
        for idx, cell in enumerate(row):
            cell.fill = PatternFill(fill_type='solid', start_color='FFFFAA', end_color='FFFFAA')
            if i == 4:
                for idx, cell in enumerate(row):
                    if idx < 4:
                        cell.border = title_med_border
                    else:
                        cell.border = med_border

    ws.freeze_panes = 'E2'
    ws.views.sheetView[0].zoomScale = 60
    # 寫出看結果
    workbook.save(file_path)
    

def set_style_for_performance(file_path:str):
    """使用openpyxl設定style。最後不會回傳return，直接寫出檔案。
    事後評估表的格式 style 都來這裡改。
    Args: 
      file_path: str. 檔案路徑。
    """
    workbook = load_workbook(file_path, data_only=True)
    ws = workbook['AD事後評估總表']
    # workbook.get_sheet_names()

    full_font = Font(name='微軟正黑體', size=12)
    full_alignment = Alignment(horizontal='center', vertical='center')
    title_font = Font(name='微軟正黑體', size=12, b=True)
    title_alignment = Alignment(horizontal='center', vertical='center')

    format_percentage = '_-* 0.00%_-;-* 0.00%_-;_-* ""??_-;_-@_-'
    format_0_decimal = '_-* #,##0_-;-* #,##0_-;_-* ""??_-;_-@_-'
    format_1_decimal = '_-* #,##0.0_-;-* #,##0.0_-;_-* ""??_-;_-@_-'
    format_2_decimal = '_-* #,##0.00_-;-* #,##0.00_-;_-* ""??_-;_-@_-'

    # 欄位設定: 寬與高
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 30.0

    ws.column_dimensions['C'].width = 20.0
    ws.column_dimensions['D'].width = 10.0

    for col in range(5, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ws.row_dimensions[1].height = 40.0
    for row in range(1, ws.max_row+1):
        ws.row_dimensions[row].height = 30.0

    # 全表設定: 字型、字體大小轉換
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            for cell in row:
                cell.font = title_font
                cell.alignment = title_alignment
        else:
            for cell in row:
                cell.font = full_font
                cell.alignment = full_alignment


    # 資料格式設定: 數值格式轉換，小數點或加上千分位逗點
    for i, col in enumerate(ws.iter_cols()):
        if i < 4:
            # 頭4欄為title
            for cell in col[1:]:
                cell.font = title_font
                cell.alignment = title_alignment
        elif ('率' in col[0].value):
            for cell in col[1:]:
                cell.number_format = format_percentage
        else:
            for cell in col[1:]:
                cell.number_format = format_0_decimal

    # 填入公式
    for i, col in enumerate(ws.iter_cols()):
    #     print(col[0].value)
        if col[0].value == '實際CPM':
            for i, cell in enumerate(col[1:]):
                cell.number_format = format_1_decimal
                cell.value = '=IFERROR(U{0}/(F{0}/1000), 0)'.format(i+2) # excel從1計算 & enumerate從0計算，所以cell要+2
        if col[0].value == '實際CPC':
            for i, cell in enumerate(col[1:]):
                cell.number_format = format_1_decimal
                cell.value = '=IFERROR(U{0}/J{0}, 0)'.format(i+2)
        if col[0].value == '實際CPV':
            for i, cell in enumerate(col[1:]):
                cell.number_format = format_2_decimal
                cell.value = '=IFERROR(U{0}/P{0}, 0)'.format(i+2)

    ws.freeze_panes = 'C2'
    ws.views.sheetView[0].zoomScale = 60
    workbook.save(file_path)